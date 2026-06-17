#!/usr/bin/env python3
"""
Attendance Report Automation System
نظام استخراج تقارير الحضور والانصراف
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime, time, date
import re

# ============================================================
# CONFIGURATION / الإعدادات
# ============================================================

SHIFT_MORNING = {
    "name": "صباحي",
    "start": time(7, 0),
    "grace": time(7, 30),
    "end": time(15, 0),
}

SHIFT_EVENING = {
    "name": "مسائي",
    "start": time(15, 0),
    "grace": time(15, 15),
    "end": time(23, 0),
}

MISSING_PUNCH = 120  # دقائق عند غياب البصمة

# ────────────────────────────────────────────────────────────
#  فلتر الفترة الزمنية — حدّد التاريخ الذي تريد التقرير منه وإليه
#  اتركهما None لعرض كل البيانات بدون فلترة
#  صيغة التاريخ: "YYYY-MM-DD"  مثال: "2026-06-01"
# ────────────────────────────────────────────────────────────
DATE_FROM = "2026-06-01"   # مثال: "2026-06-01"
DATE_TO   = "2026-06-30"   # مثال: "2026-06-30"

# ============================================================
# HELPERS
# ============================================================

def parse_time(val):
    if pd.isna(val):
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(val, fmt).time()
            except ValueError:
                pass
    return None


def minutes_between(t1, t2) -> int:
    d = datetime(2000, 1, 1)
    return int((datetime.combine(d, t2) - datetime.combine(d, t1)).total_seconds() // 60)


def calc_late(in_time, shift: dict) -> int:
    if in_time is None:
        return MISSING_PUNCH
    if in_time <= shift["grace"]:
        return 0
    return minutes_between(shift["start"], in_time)


def calc_early_leave(out_time, shift: dict) -> int:
    if out_time is None:
        return MISSING_PUNCH
    if out_time >= shift["end"]:
        return 0
    return minutes_between(out_time, shift["end"])


def detect_shift(in_time, out_time) -> dict:
    if in_time is None:
        if out_time and out_time >= time(15, 0):
            return SHIFT_EVENING
        return SHIFT_MORNING
    if in_time >= time(14, 0):
        return SHIFT_EVENING
    return SHIFT_MORNING


def parse_date_any(val):
    """Parse a date value (str or datetime) into a date object."""
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        try:
            return pd.to_datetime(val).date()
        except Exception:
            return None
    try:
        return pd.to_datetime(val).date()
    except Exception:
        return None


def in_date_range(d, date_from, date_to) -> bool:
    if d is None:
        return False
    if date_from and d < date_from:
        return False
    if date_to and d > date_to:
        return False
    return True


def fmt_date_excel(d) -> str:
    """Format as M/D/YYYY to match reference output."""
    if d is None:
        return ""
    return f"{d.month}/{d.day}/{d.year}"


def time_str(t) -> str:
    if t is None:
        return ""
    # تنسيق 12 ساعة مع AM/PM (مثال: 7:30 AM)
    return t.strftime("%I:%M %p").lstrip("0")


# ============================================================
# EXCEL STYLE HELPERS
# ============================================================

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def header_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    return cell


def data_cell(ws, row, col, value, bg=None, bold=False, color=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", size=10, bold=bold, color=color if color else "000000")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    return cell


# ============================================================
# LOAD DATA
# ============================================================

def load_data(filepath: str) -> pd.DataFrame:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(filepath, encoding="utf-8-sig")
    else:
        df = pd.read_excel(filepath)

    col_map = {}
    used_targets = set()
    for c in df.columns:
        cl = str(c).lower().strip()
        target = None
        if "name" in cl or "اسم" in cl:
            target = "name"
        elif "date" in cl or "تاريخ" in cl:
            target = "date"
        elif ("check in" in cl) or ("clock in" in cl) or cl.endswith(" in 1") or cl.endswith(" in") or "دخول" in cl or "حضور" in cl:
            target = "in_time"
        elif ("check out" in cl) or ("clock out" in cl) or cl.endswith(" out 1") or cl.endswith(" out") or "خروج" in cl or "انصراف" in cl:
            target = "out_time"
        # تجنّب تكرار نفس الهدف لعمود ثانٍ (مثل أعمدة دقائق التأخير التالفة الإكسل)
        if target and target not in used_targets:
            col_map[c] = target
            used_targets.add(target)

    df.rename(columns=col_map, inplace=True)
    required = {"name", "date", "in_time", "out_time"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"أعمدة مفقودة: {missing}\nالأعمدة الموجودة: {list(df.columns)}")
    return df


def process_employee(records: list) -> list:
    # نخزّن السجلات الفعلية حسب التاريخ، ثم نملأ كل أيام الفترة المحددة
    # (إذا كانت محددة) — أي يوم بدون بصمة يظهر بتأخير=120 وخروج مبكر=120
    by_date = {}
    for r in records:
        d = parse_date_any(r.get("date"))
        if d is None:
            continue
        in_t = parse_time(r.get("in_time"))
        out_t = parse_time(r.get("out_time"))
        by_date[d] = (in_t, out_t)

    if DATE_FROM_PARSED and DATE_TO_PARSED:
        # نولّد كل تواريخ الفترة بالكامل (حتى الأيام بدون أي بصمة)
        all_dates = []
        d = DATE_FROM_PARSED
        one_day = pd.Timedelta(days=1)
        while d <= DATE_TO_PARSED:
            all_dates.append(d)
            d = (pd.Timestamp(d) + one_day).date()
    else:
        # بدون فلتر فترة: نعرض فقط الأيام التي فيها سجل فعلي (كالسابق)
        all_dates = sorted(by_date.keys())

    rows = []
    for d in all_dates:
        if DATE_FROM_PARSED or DATE_TO_PARSED:
            if not in_date_range(d, DATE_FROM_PARSED, DATE_TO_PARSED):
                continue
        in_t, out_t = by_date.get(d, (None, None))
        shift = detect_shift(in_t, out_t)
        late = calc_late(in_t, shift)
        early = calc_early_leave(out_t, shift)
        rows.append({
            "date": d,
            "in_time": in_t,
            "out_time": out_t,
            "late": late,
            "early": early,
        })
    rows.sort(key=lambda r: r["date"])
    return rows


# ============================================================
# GENERATE REPORT — matches reference layout exactly
# Columns: Name | Date | Check In | Check Out | Late(min) | Early(min)
# ============================================================

def generate_report(employee_name: str, rows: list, output_dir: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.sheet_view.rightToLeft = True

    headers = ["الاسم", "التاريخ", "الدخول", "الخروج", "التأخير (دقيقة)", "الخروج المبكر (دقيقة)"]
    col_widths = [22, 14, 12, 12, 16, 18]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    for col, h in enumerate(headers, 1):
        header_cell(ws, 1, col, h)
    ws.row_dimensions[1].height = 26

    r = 2
    for row in rows:
        late = row["late"]
        early = row["early"]

        if late > 0 and early > 0:
            bg = "FFD5D5"
        elif late > 0:
            bg = "FFE0E0"
        elif early > 0:
            bg = "FFF3CD"
        else:
            bg = None

        data_cell(ws, r, 1, employee_name.strip(), bg=bg)
        data_cell(ws, r, 2, fmt_date_excel(row["date"]), bg=bg)
        data_cell(ws, r, 3, time_str(row["in_time"]), bg=bg,
                  color="C00000" if row["in_time"] is None else None)
        data_cell(ws, r, 4, time_str(row["out_time"]), bg=bg,
                  color="C00000" if row["out_time"] is None else None)
        data_cell(ws, r, 5, late, bg=bg, bold=late > 0, color="C00000" if late > 0 else "375623")
        data_cell(ws, r, 6, early, bg=bg, bold=early > 0, color="ED7D31" if early > 0 else "375623")
        r += 1

    # ---- صف المجموع: E و F كل واحد بمعادلة SUM الخاصة به ----
    total_row = r
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    data_cell(ws, total_row, 1, "المجموع", bold=True, bg="F2F2F2")

    if len(rows) > 0:
        late_formula = f"=SUM(E2:E{r - 1})"
        early_formula = f"=SUM(F2:F{r - 1})"
    else:
        late_formula = 0
        early_formula = 0

    data_cell(ws, total_row, 5, late_formula, bold=True, bg="F2F2F2")
    data_cell(ws, total_row, 6, early_formula, bold=True, bg="F2F2F2")

    # ---- صف إضافي تحت المجموع: مجموع المجموعين معاً (E33+F33) ----
    grand_row = total_row + 1
    ws.merge_cells(start_row=grand_row, start_column=1, end_row=grand_row, end_column=4)
    data_cell(ws, grand_row, 1, "إجمالي (تأخير + خروج مبكر)", bold=True, bg="DCE6F1", color="1F4E79")

    grand_formula = f"=E{total_row}+F{total_row}"
    ws.merge_cells(start_row=grand_row, start_column=5, end_row=grand_row, end_column=6)
    data_cell(ws, grand_row, 5, grand_formula, bold=True, bg="DCE6F1", color="1F4E79")

    ws.freeze_panes = "A2"
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    safe_name = re.sub(r'[\\/*?:"<>|]', '_', employee_name.strip())
    out_path = os.path.join(output_dir, f"{safe_name}.xlsx")
    wb.save(out_path)
    return out_path


# ============================================================
# MAIN
# ============================================================

DATE_FROM_PARSED = None
DATE_TO_PARSED = None


def main():
    global DATE_FROM_PARSED, DATE_TO_PARSED
    DATE_FROM_PARSED = parse_date_any(DATE_FROM) if DATE_FROM else None
    DATE_TO_PARSED = parse_date_any(DATE_TO) if DATE_TO else None

    base = os.path.dirname(os.path.abspath(__file__))
    input_dir = os.path.join(base, "input")
    output_dir = os.path.join(base, "output")
    os.makedirs(output_dir, exist_ok=True)

    csv_files = [f for f in os.listdir(input_dir) if f.endswith(".csv")]
    xlsx_files = [f for f in os.listdir(input_dir) if f.endswith(".xlsx")]
    all_files = csv_files + xlsx_files
    if not all_files:
        print("❌ لم يتم العثور على ملف بيانات في مجلد input")
        sys.exit(1)

    input_file = os.path.join(input_dir, all_files[0])
    print(f"📂 قراءة الملف: {all_files[0]}")

    if DATE_FROM_PARSED or DATE_TO_PARSED:
        print(f"📅 فلتر الفترة: من {DATE_FROM_PARSED or '—'} إلى {DATE_TO_PARSED or '—'}")

    df = load_data(input_file)
    df = df[df["name"].notna()]
    df = df[~df["name"].astype(str).str.strip().isin(["المجموع", "Total", "الإجمالي", ""])]

    employees = df.groupby("name")
    print(f"👥 عدد الموظفين: {len(employees)}")

    generated = []
    for name, group in employees:
        name = str(name).strip()
        records = group.to_dict("records")
        rows = process_employee(records)
        if not rows:
            print(f"  ⚠️  {name}: لا توجد بيانات ضمن الفترة المحددة")
            continue
        path = generate_report(name, rows, output_dir)
        total_late = sum(r["late"] for r in rows)
        total_early = sum(r["early"] for r in rows)
        total_all = total_late + total_early
        print(f"  ✅ {name} → {len(rows)} يوم | تأخير: {total_late}د | خروج مبكر: {total_early}د | الإجمالي: {total_all}د")
        generated.append(path)

    print(f"\n✅ تم إنشاء {len(generated)} تقرير في مجلد output/")
    return generated


if __name__ == "__main__":
    main()
